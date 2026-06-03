import ast
import os
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

os.environ.setdefault("MPLBACKEND", "Agg")

import pandas as pd
from openpyxl import load_workbook
from rich.console import Console
from rich.panel import Panel
from rich.prompt import Confirm, IntPrompt, Prompt
from rich.table import Table

from src.preprocessing.utils.normalization import NormalizerType
from src.preprocessing.utils.sampling import SamplerType
from src.preprocessing.utils.target import TargetType


console = Console()


class CLIError(ValueError):
    """User-facing CLI error."""


TARGET_OPTIONS = {
    "result": TargetType.RESULT,
    "1x2": TargetType.RESULT,
    "over-under": TargetType.OVER_UNDER,
    "over_under": TargetType.OVER_UNDER,
    "uo": TargetType.OVER_UNDER,
    "u/o": TargetType.OVER_UNDER,
}

TARGET_LABELS = {
    TargetType.RESULT: "Result (1/X/2)",
    TargetType.OVER_UNDER: "U/O-2.5",
}

NORMALIZER_OPTIONS = {
    "none": None,
    "standard": NormalizerType.STANDARD,
    "min-max": NormalizerType.MIN_MAX,
    "minmax": NormalizerType.MIN_MAX,
    "max-abs": NormalizerType.MAX_ABS,
    "maxabs": NormalizerType.MAX_ABS,
}

SAMPLER_OPTIONS = {
    "none": None,
    "svm-smote": SamplerType.SVM_SMOTE,
    "svm_smote": SamplerType.SVM_SMOTE,
    "nearmiss": SamplerType.NEARMISS,
    "near-miss": SamplerType.NEARMISS,
    "hardness-threshold": SamplerType.INSTANCE_HARDNESS_THRESHOLD,
    "instance-hardness-threshold": SamplerType.INSTANCE_HARDNESS_THRESHOLD,
}

COLORMAP_OPTIONS = {
    "Blues": "Blues",
    "Coolwarm": "coolwarm",
    "Crest": "crest",
    "HUSL": "husl",
    "Icefire": "icefire",
    "Rocket": "rocket",
    "Summer": "summer",
}

HELP_LINKS = {
    "Machine Learning": "https://www.ibm.com/think/topics/machine-learning",
    "ML vs Statistics": "https://www.geeksforgeeks.org/machine-learning/difference-between-statistical-model-and-machine-learning/",
    "Supervised Learning": "https://www.geeksforgeeks.org/machine-learning/supervised-machine-learning/",
    "Classification": "https://www.geeksforgeeks.org/machine-learning/getting-started-with-classification/",
    "Classification Metrics": "https://www.geeksforgeeks.org/machine-learning/metrics-for-machine-learning-model/",
    "KNN": "https://www.ibm.com/think/topics/knn",
    "Naive Bayes": "https://www.geeksforgeeks.org/machine-learning/naive-bayes-classifiers/",
    "Logistic Regression": "https://www.ibm.com/think/topics/logistic-regression",
    "Decision Tree": "https://www.ibm.com/think/topics/decision-trees",
    "Random Forest": "https://builtin.com/data-science/random-forest-algorithm",
    "XGBoost": "https://www.ibm.com/think/topics/xgboost",
    "LDA": "https://www.geeksforgeeks.org/machine-learning/ml-linear-discriminant-analysis/",
    "SVM": "https://spotintelligence.com/2024/05/06/support-vector-machines-svm/",
    "DNN": "https://www.geeksforgeeks.org/machine-learning/neural-networks-a-beginners-guide/",
    "Feature Normalization": "https://www.geeksforgeeks.org/machine-learning/Feature-Engineering-Scaling-Normalization-and-Standardization/",
    "DNN Improvements": "https://stanford.edu/~shervine/teaching/cs-230/cheatsheet-deep-learning-tips-and-tricks",
    "Interpretability": "https://christophm.github.io/interpretable-ml-book/interpretability.html",
    "Variable Distribution": "https://www.linkedin.com/pulse/understanding-data-distributions-examples-applications-r-s-8gkhc/",
    "Feature Correlations": "https://users.sussex.ac.uk/~grahamh/RM1web/Eight%20things%20you%20need%20to%20know%20about%20interpreting%20correlations.pdf",
    "Feature Variance": "https://www.geeksforgeeks.org/maths/variance/",
    "Regression Coefficients": "https://articles.outlier.org/coefficient-regression",
    "Impurity": "https://www.geeksforgeeks.org/machine-learning/gini-impurity-and-entropy-in-decision-tree-ml/",
    "Boruta": "https://www.blog.trainindata.com/is-boruta-dead/",
    "Class Imbalance": "https://isi-web.org/sites/default/files/2024-02/Handling-Data-Imbalance-in-Machine-Learning.pdf",
    "SMOTE": "https://www.analyticsvidhya.com/blog/2020/10/overcoming-class-imbalance-using-smote-techniques/",
    "NearMiss": "https://www.linkedin.com/pulse/under-sampling-method-kaamil-ahmed/",
    "Partial Dependence Plot": "https://www.geeksforgeeks.org/deep-learning/partial-dependence-plot-from-an-xgboost-model-in-r/",
    "SHAP": "https://shap.readthedocs.io/en/latest/example_notebooks/overviews/An%20introduction%20to%20explainable%20AI%20with%20Shapley%20values.html",
    "Update": "https://github.com/kochlisGit/ProphitBet-Soccer-Bets-Predictor/tree/main",
    "Submit Bug": "https://github.com/kochlisGit/ProphitBet-Soccer-Bets-Predictor/issues/new",
    "Donation": "https://www.paypal.com/donate/?hosted_button_id=AK3SEFDGVAWFE",
}


def normalize_key(value: str) -> str:
    return value.strip().lower().replace("_", "-")


def parse_target(value: str) -> TargetType:
    key = normalize_key(value)
    if key not in TARGET_OPTIONS:
        raise CLIError(f'Invalid target "{value}". Use: result, over-under.')
    return TARGET_OPTIONS[key]


def parse_normalizer(value: str):
    key = normalize_key(value)
    if key not in NORMALIZER_OPTIONS:
        raise CLIError(f'Invalid normalizer "{value}". Use: none, standard, min-max, max-abs.')
    return NORMALIZER_OPTIONS[key]


def parse_sampler(value: str):
    key = normalize_key(value)
    if key not in SAMPLER_OPTIONS:
        raise CLIError(
            f'Invalid sampler "{value}". Use: none, svm-smote, nearmiss, hardness-threshold.'
        )
    return SAMPLER_OPTIONS[key]


def target_label(target_type: TargetType) -> str:
    return TARGET_LABELS[target_type]


def validate_identifier(value: str, label: str = "identifier") -> str:
    value = value.strip()
    if not value:
        raise CLIError(f"{label} cannot be empty.")
    if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9_.-]*", value):
        raise CLIError(
            f'{label} "{value}" is invalid. Use letters, numbers, dots, underscores or hyphens.'
        )
    return value


def parse_columns(value: Optional[str]) -> Optional[List[str]]:
    if value is None:
        return None
    columns = [col.strip() for col in value.split(",") if col.strip()]
    return columns or None


def parse_odd_range(value: Optional[str]) -> Optional[Tuple[float, float]]:
    if value is None or str(value).strip().lower() in {"", "none", "all"}:
        return None

    raw = str(value).strip().replace(",", ":")
    parts = [part.strip() for part in raw.split(":") if part.strip()]
    if len(parts) != 2:
        raise CLIError(f'Invalid odd range "{value}". Use MIN:MAX, for example 1.30:2.50.')

    try:
        low, high = float(parts[0]), float(parts[1])
    except ValueError as exc:
        raise CLIError(f'Invalid odd range "{value}". Both bounds must be numeric.') from exc

    if low < 1.0 or high < 1.0:
        raise CLIError("Odd ranges must be greater than or equal to 1.0.")
    if low == 1.0 and high == 10.0:
        return None
    if high <= low:
        raise CLIError("Odd range max must be greater than min.")
    if high - low < 0.5:
        raise CLIError("Odd range max-min difference must be at least 0.5.")
    if high == 10.0:
        high = 1000.0
    return low, high


def parse_eval_odd_range(value: Optional[str]):
    if value is None or str(value).strip().lower() in {"", "none"}:
        return "None"

    raw = str(value).strip()
    if raw.startswith("("):
        parsed = ast.literal_eval(raw)
        if parsed == "None":
            return "None"
        if isinstance(parsed, tuple) and len(parsed) == 3:
            odd, low, high = parsed
            return str(odd), float(low), float(high)
        raise CLIError(f'Invalid stored odd filter "{value}".')

    parts = [part.strip() for part in raw.replace(",", ":").split(":") if part.strip()]
    if len(parts) != 3:
        raise CLIError('Evaluation odd filter must be "ODD:MIN:MAX", for example "1:1.31:1.60".')

    odd, low, high = parts[0], parts[1], parts[2]
    if odd not in {"1", "X", "2"}:
        raise CLIError('Evaluation odd filter ODD must be one of "1", "X", "2".')

    try:
        return odd, float(low), float(high)
    except ValueError as exc:
        raise CLIError(f'Invalid odd filter "{value}". Bounds must be numeric.') from exc


def parse_bool(value: str) -> bool:
    key = normalize_key(value)
    if key in {"1", "true", "yes", "y", "si", "sí", "on"}:
        return True
    if key in {"0", "false", "no", "n", "off"}:
        return False
    raise CLIError(f'Invalid boolean value "{value}".')


def parse_tunable_params(value: Optional[str]) -> List[str]:
    if value is None or value.strip().lower() in {"", "none"}:
        return []
    if value.strip().lower() == "all":
        return ["all"]
    return [param.strip() for param in value.split(",") if param.strip()]


def confirm_or_abort(message: str, assume_yes: bool = False):
    if assume_yes:
        return
    if not Confirm.ask(message, default=False):
        raise CLIError("Operation cancelled.")


def prompt_choice(
        title: str,
        choices: Sequence[Any],
        label_fn=str,
        default_index: int = 0,
) -> Any:
    if not choices:
        raise CLIError(f"No choices available for {title}.")

    table = Table(title=title, show_lines=False)
    table.add_column("#", justify="right")
    table.add_column("Option")
    for idx, choice in enumerate(choices, start=1):
        table.add_row(str(idx), label_fn(choice))
    console.print(table)
    selected = IntPrompt.ask(
        "Select option",
        default=default_index + 1,
        choices=[str(i) for i in range(1, len(choices) + 1)],
    )
    return choices[selected - 1]


def prompt_text(message: str, default: Optional[str] = None) -> str:
    if default is None:
        return Prompt.ask(message)
    return Prompt.ask(message, default=default)


def prompt_int(message: str, default: int, min_value: Optional[int] = None, max_value: Optional[int] = None) -> int:
    while True:
        value = IntPrompt.ask(message, default=default)
        if min_value is not None and value < min_value:
            console.print(f"[red]Value must be >= {min_value}.[/red]")
            continue
        if max_value is not None and value > max_value:
            console.print(f"[red]Value must be <= {max_value}.[/red]")
            continue
        return value


def render_dataframe(
        df: pd.DataFrame,
        title: str,
        max_rows: int = 20,
        columns: Optional[Sequence[str]] = None,
        show_index: bool = False,
):
    if columns is not None:
        missing = [col for col in columns if col not in df.columns]
        if missing:
            raise CLIError(f"Unknown columns: {', '.join(missing)}")
        df = df[list(columns)]

    display_df = df.head(max_rows)
    table = Table(title=title, show_lines=False)
    if show_index:
        table.add_column("#", justify="right")
    for column in display_df.columns:
        table.add_column(str(column), overflow="fold")

    for idx, row in display_df.iterrows():
        values = ["" if pd.isna(value) else str(value) for value in row.tolist()]
        if show_index:
            values.insert(0, str(idx))
        table.add_row(*values)

    console.print(table)
    if df.shape[0] > display_df.shape[0]:
        console.print(f"[dim]Showing {display_df.shape[0]} of {df.shape[0]} rows.[/dim]")


def render_mapping(title: str, values: Dict[str, Any]):
    table = Table(title=title)
    table.add_column("Key")
    table.add_column("Value")
    for key, value in values.items():
        table.add_row(str(key), str(value))
    console.print(table)


def ensure_output_path(path: Optional[str], default_name: str, suffix: str) -> Path:
    output = Path(path) if path else Path("outputs") / default_name
    if output.suffix == "":
        output = output.with_suffix(suffix)
    output.parent.mkdir(parents=True, exist_ok=True)
    return output


def export_dataframe(df: pd.DataFrame, path: str, append: bool = False):
    output = Path(path)
    if output.suffix == "":
        output = output.with_suffix(".csv")
    output.parent.mkdir(parents=True, exist_ok=True)

    suffix = output.suffix.lower()
    if suffix == ".csv":
        mode = "a" if append and output.exists() else "w"
        header = not (append and output.exists())
        df.to_csv(output, mode=mode, header=header, index=False)
    elif suffix == ".xlsx":
        if append and output.exists():
            workbook = load_workbook(filename=output)
            sheet_name = "Sheet1"
            startrow = workbook[sheet_name].max_row if sheet_name in workbook.sheetnames else 0
            mode, header, if_sheet_exists = "a", False, "overlay"
        else:
            startrow, mode, header, if_sheet_exists = 0, "w", True, None
        with pd.ExcelWriter(output, engine="openpyxl", mode=mode, if_sheet_exists=if_sheet_exists) as writer:
            df.to_excel(writer, sheet_name="Sheet1", startrow=startrow, header=header, index=False)
    else:
        raise CLIError(f'Unsupported output format "{suffix}". Use .csv or .xlsx.')

    console.print(f"[green]Exported:[/green] {output}")


def save_figure(ax, path: str):
    import matplotlib.pyplot as plt
    import numpy as np

    output = Path(path)
    if output.suffix == "":
        output = output.with_suffix(".png")
    output.parent.mkdir(parents=True, exist_ok=True)

    if isinstance(ax, np.ndarray):
        fig = ax.ravel()[0].figure
    elif isinstance(ax, list):
        fig = ax[0].figure
    else:
        fig = ax.figure
    fig.savefig(output, bbox_inches="tight", dpi=160)
    plt.close(fig)
    console.print(f"[green]Saved figure:[/green] {output}")


def load_required_columns(df: pd.DataFrame, required: Iterable[str], context: str):
    missing = [column for column in required if column not in df.columns]
    if missing:
        raise CLIError(f"{context} is missing required columns: {', '.join(missing)}")


def print_success(message: str):
    console.print(Panel(message, style="green", title="Done"))


def print_warning(message: str):
    console.print(Panel(message, style="yellow", title="Warning"))


def print_error(message: str):
    console.print(Panel(message, style="red", title="Error"))
